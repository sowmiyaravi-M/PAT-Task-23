import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime

class OrangeHRMLoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_input = (By.ID, 'txtUsername')
        self.password_input = (By.ID, 'txtPassword')
        self.login_button = (By.ID, 'btnLogin')

    def login(self, username, password):
        self.driver.find_element(*self.username_input).clear()
        self.driver.find_element(*self.username_input).send_keys(username)
        self.driver.find_element(*self.password_input).clear()
        self.driver.find_element(*self.password_input).send_keys(password)
        self.driver.find_element(*self.login_button).click()


@pytest.fixture
def browser():
    driver = webdriver.Chrome()
    driver.implicitly_wait(10)
    yield driver
    driver.quit()


def test_login(browser):
    workbook = load_workbook('test_data.xlsx')
    sheet = workbook.active
    row_count = sheet.max_row
    for i in range(2, row_count + 1):
        username = sheet.cell(row=i, column=2).value
        password = sheet.cell(row=i, column=3).value
        driver = browser
        driver.get('https://opensource-demo.orangehrmlive.com/web/index.php/auth/login')
        login_page = OrangeHRMLoginPage(driver)
        login_page.login(username, password)
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'welcome'))
            )
            sheet.cell(row=i, column=4).value = datetime.now().strftime('%Y-%m-%d')
            sheet.cell(row=i, column=5).value = datetime.now().strftime('%H:%M:%S')
            sheet.cell(row=i, column=6).value = 'Test Passed'
        except:
            sheet.cell(row=i, column=6).value = 'Test Failed'
        workbook.save('test_data.xlsx')


if __name__ == "__main__":
    pytest.main(['-v', 'test_script.py'])
